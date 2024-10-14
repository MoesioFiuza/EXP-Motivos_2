import pandas as pd
from openpyxl import load_workbook

planilha = r'C:\Users\moesios\Desktop\AJUSTES P14\matrizes.xlsx'

df_viagens = pd.read_excel(planilha, sheet_name='Planilha1')

motivos_desejados = ['OUTROS', 'TRABALHO', 'ESTUDOS', 'NÃO DOMICILIAR']

viagens_filtradas = df_viagens[df_viagens['MOTIVO DA VIAGEM'].isin(motivos_desejados)]

contagem_viagens = viagens_filtradas.groupby(['ZONA ORIGEM', 'ZONA DESTINO']).size().reset_index(name='count')

def get_fator_expansao(row):
    filtered = df_viagens[(df_viagens['ZONA ORIGEM'] == row['ZONA ORIGEM']) & 
                          (df_viagens['ZONA DESTINO'] == row['ZONA DESTINO']) & 
                          (df_viagens['MOTIVO DA VIAGEM'] == 'ESTUDOS')]
    if not filtered.empty:
        return filtered['FATOR DE EXPANSÃO'].mean()
    return 1

contagem_viagens['FATOR_EXPANSAO'] = contagem_viagens.apply(get_fator_expansao, axis=1)

contagem_viagens['count'] *= contagem_viagens['FATOR_EXPANSAO']

for motivo in motivos_desejados:
    contagem_viagens[motivo] = contagem_viagens.apply(lambda row: len(viagens_filtradas[(viagens_filtradas['ZONA ORIGEM'] == row['ZONA ORIGEM']) & 
                                                                                         (viagens_filtradas['ZONA DESTINO'] == row['ZONA DESTINO']) & 
                                                                                         (viagens_filtradas['MOTIVO DA VIAGEM'] == motivo)]), axis=1)

total_expansao = contagem_viagens['count'].sum()
print(f"Total da expansão: {total_expansao}")

book = load_workbook(planilha)
if 'FUMO saida' in book.sheetnames:
    del book['FUMO saida']

with pd.ExcelWriter(planilha, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    contagem_viagens.to_excel(writer, sheet_name='FUMO saida', index=False)